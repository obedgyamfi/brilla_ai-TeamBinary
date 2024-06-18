import openpyxl
import re
import json

def extract_latex_expressions(file_path):
    expressions_with_locations = []

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        if sheet_name.startswith('Round'):
            # Iterate through columns C, D, and I
            for row in range(1, sheet.max_row + 1):
                for col in ['C', 'D', 'I']:
                    cell = sheet[f'{col}{row}']
                    if isinstance(cell.value, str):
                        latex_expressions = re.findall(r'\$([^$]+)\$', cell.value)
                        if latex_expressions:
                            for latex_expression in latex_expressions:
                                expression_location = {
                                    'expression': latex_expression,
                                    'sheet': sheet_name,
                                    'row': row,
                                    'column': col,
                                    'sentence': ''  # Adding an empty 'sentence' key
                                }
                                expressions_with_locations.append(expression_location)

    return expressions_with_locations

# Example usage
file_path = '../ai_models_and_data/models_and_data/nsmq_past_questions/NSMQ QUESTIONS SPREADSHEETS/2020/2020 NSMQ contest 1.xlsx'  # Replace with the actual file path
expressions_with_locations = extract_latex_expressions(file_path)

# Convert the result to JSON
json_output = json.dumps(expressions_with_locations, indent=4)

# Store the JSON output in a text file
with open('expressions_with_locations.json', 'w') as json_file:
    json_file.write(json_output)

