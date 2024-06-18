import os
import openpyxl
import re

# Function to extract LaTeX expressions from a string
def extract_latex(text):
    latex_expressions = re.findall(r'\$([^$]+)\$', text)
    return latex_expressions

# Function to process each Excel file in the directory
def process_excel_files(directory_path, output_file):
    with open(output_file, 'w') as out_file:
        for filename in os.listdir(directory_path):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(directory_path, filename)
                workbook = openpyxl.load_workbook(file_path)
                print(f'Processing file: {filename}')
                # Iterate through each sheet in the workbook
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    if sheet_name.startswith('Round'):
                        print(f'Processing sheet: {sheet_name}')
                        # Iterate through columns C, D, and I
                        for row in range(1, sheet.max_row + 1):
                            for col in ['C', 'D', 'I']:
                                cell = sheet[f'{col}{row}']
                                if isinstance(cell.value, str):
                                    latex_expressions = extract_latex(cell.value)
                                    if latex_expressions:
                                        for latex_expression in latex_expressions:
                                            out_file.write(f'{latex_expression}\n')
                print(f'Processed file: {filename}')

# Directory path containing the Excel files
directory_path = '../ai_models_and_data/models_and_data/nsmq_past_questions/NSMQ QUESTIONS SPREADSHEETS/2021/'  # Update with the directory containing your Excel files
output_file = './latex_expressions.txt'  # Update with the path to the output file
process_excel_files(directory_path, output_file)

