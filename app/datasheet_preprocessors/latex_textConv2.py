import os
import openpyxl
import re
import shutil

# Conversion dictionary
conversion_dict = {
    r'\\frac\{(-?\d+)\}\{(-?\d+)\}': lambda m: f"{m.group(1)} over {m.group(2)}",
    r'(-?\d+)\s*x\^\{(\d+)\}': r'\1 times x to the power of \2',
    r'(-?\d+)\s*x': r'\1 times x',
    r'x\^\{(\d+)\}': r'x to the power of \1',
    r'\\left': '',
    r'\\right': '',
    r'\\mathrm': '',
    r'\\begin': '',
    r'\\text': '',
    r'\\end': '',
    r'\\quad': ' ',
    r'\\equiv': ' is equivalent to ',
    r'\\cos': 'cosine ',
    r'\\sin': 'sine ',
    r'\\sqrt': 'square root of ',
    r'\\times': ' times ',
    r'\\mu': 'mu ',
    r'\\pi': 'pi ',
    r'\\rightarrow': 'rightarrow ',
    r'\\Omega': 'Omega ',
    r'\\mathbf': '',
    r'\\cdot': ' dot ',
    r'\\hat': 'hat ',
    r'\\circ': 'circle ',
    r'\\tan': 'tangent ',
    r'\\Delta': 'Delta ',
    r'\\int': 'integral ',
    r'\\infty': 'infinity ',
    r'\\Pi': 'Pi ',
    r'\\leq': ' less than or equal to ',
    r'\\theta': 'theta ',
    r'\\lambda': 'lambda ',
    r'\\alpha': 'alpha ',
    r'\\pm': ' plus or minus ',
    r'\\neq': ' not equal to ',
    r'\\rightleftharpoons': ' rightleftharpoons ',
    r'\\log': 'logarithm ',
    r'\\cong': ' congruent to ',
    r'\\eta': 'eta ',
    r'\\cot': 'cotangent ',
    r'\\min': ' minimum ',
    r'\\geq': ' greater than or equal to ',
    r'\\operatorname': '',
    r'\\underline': 'underline ',
    r'\\beta': 'beta ',
    r'\\chi': 'chi ',
    r'\\epsilon': 'epsilon ',
    r'\\rho': 'rho ',
    r'\\propto': 'proportional to ',
    r'\\Sigma': 'Sigma ',
    r'\\ldots': '...',
    r'\\leftrightarrow': 'leftrightarrow ',
    r'\\sec': 'secant ',
    r'\\uparrow': 'uparrow ',
    r'\\cup': 'union ',
    r'\\cap': 'intersection ',
    r'\\div': ' divided by ',
    r'(-?\d+)': r'\1',
    r'\+': ' plus ',
    r'-': ' minus ',
    r'=': ' equals ',
    # Units and Constants
    r'\mathrm\{J\}': 'Joules',
    r'\mathrm\{Pa\}': 'Pascals',
    r'\mathrm\{m\}': 'meters',
    r'\mathrm\{N\}': 'Newtons',
    r'\mathrm\{kg\}': 'kilograms',
    r'\mathrm\{g\}': 'grams',
    r'\mathrm\{s\}': 'seconds',
    r'\mathrm\{A\}': 'Amperes',
    r'\mathrm\{F\}': 'Farads',
    r'\mathrm\{V\}': 'Volts',
    r'\mathrm\{C\}': 'Coulombs',
    r'\mathrm\{ohm\}': 'ohms',
    r'\mathrm\{Hz\}': 'Hertz',
    r'\mathrm\{K\}': 'Kelvin',
    r'\mathrm\{mol\}': 'moles',
    r'\mathrm\{W\}': 'Watts',
    r'\mathrm\{H\}': 'Henries',
    r'\mu\mathrm\{m\}': 'micrometers',
    r'\mu\mathrm\{F\}': 'microfarads',
    r'\mu\mathrm\{s\}': 'microseconds',
    r'\mathrm\{cm\}': 'centimeters',
    r'\mathrm\{mm\}': 'millimeters',
    r'\mathrm\{km\}': 'kilometers',
    r'\mathrm\{dm\}': 'decimeters',
    r'\mathrm\{pm\}': 'picometers',
    r'\mathrm\{ms\}': 'milliseconds',
    r'\mathrm\{nm\}': 'nanometers',
    r'\mathrm\{ms\}^{-1}': 'meters per second',
    r'\mathrm\{kg\} \cdot \mathrm\{m\} \cdot \mathrm\{s\}^{-2}': 'Newtons',
    r'\mathrm\{m\} \cdot \mathrm\{s\}^{-2}': 'acceleration',
    r'\mathrm\{m\}^{-2}': 'per square meter',
    r'\mathrm\{m\}^{3}': 'cubic meters',
    r'\mathrm\{kg\} \cdot \mathrm\{m\}^{-3}': 'density',
}

# Function to convert LaTeX expression to readable text using the conversion dictionary
def latex_to_readable(latex_expression):
    # Remove the surrounding $ symbols
    latex_expression = latex_expression.strip('$')
    
    # Apply conversions
    readable_text = latex_expression
    for pattern, replacement in conversion_dict.items():
        if callable(replacement):
            readable_text = re.sub(pattern, replacement, readable_text)
        else:
            readable_text = re.sub(pattern, replacement, readable_text)
    
    return readable_text.strip()

# Function to process each Excel file in the directory
def process_excel_files(directory_path, output_directory):
    # Ensure the output directory exists
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    for filename in os.listdir(directory_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(directory_path, filename)
            # Copy the original file to a new file in the output directory
            copied_file_path = os.path.join(output_directory, filename)
            shutil.copyfile(file_path, copied_file_path)
            
            # Load the copied workbook
            workbook = openpyxl.load_workbook(copied_file_path)
            print(f'Processing file: {filename}')
            
            # Iterate through each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet_name.startswith('Round'):
                    print(f'Processing sheet: {sheet_name}')
                    # Iterate through columns C, D, and I
                    for col in ['C', 'D', 'I']:
                        for row in range(1, sheet.max_row + 1):
                            cell = sheet[f'{col}{row}']
                            if isinstance(cell.value, str):
                                latex_expressions = re.findall(r'\$.*?\$', cell.value)
                                for latex_expression in latex_expressions:
                                    readable_text = latex_to_readable(latex_expression)
                                    print(f'Row {row} Col {col} LaTeX: {latex_expression} -> Readable: {readable_text}')
                                    cell.value = cell.value.replace(latex_expression, readable_text)
            # Save the modified workbook
            workbook.save(copied_file_path)
            print(f'Modified file saved as {copied_file_path}')

# Directory path containing the Excel files
directory_path = '../ai_models_and_data/models_and_data/nsmq_past_questions/NSMQ QUESTIONS SPREADSHEETS/2020/'  # Update with the directory containing your Excel files
output_directory = '../ai_models_and_data/models_and_data/nsmq_past_questions/NSMQ QUESTIONS SPREADSHEETS/2020/modified_files/'  # Directory to save modified files
process_excel_files(directory_path, output_directory)
