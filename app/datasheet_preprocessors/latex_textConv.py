import openpyxl
import re

# Conversion dictionary
'''
conversion_dict = {
    r'\\frac\{(-?\d+)\}\{(-?\d+)\}': lambda m: f"{m.group(1)} over {m.group(2)}",
    r'(-?\d+)\s*x\^\{(\d+)\}': r'\1 times x to the power of \2',
    r'(-?\d+)\s*x': r'\1 times x',
    r'x\^\{(\d+)\}': r'x to the power of \1',
    r'(-?\d+)': r'\1',
    r'\+': ' plus ',
    r'-': ' minus ',
    r'=': ' equals ',
}
'''
conversion_dict = {
    r'\\ldots': 'and so on',
    r'\\frac\{\\sqrt\{(.*?)\}\}\{\\sqrt\{(.*?)\}\}': lambda m: fr"the square root of {m.group(1)} over the square root of {m.group(2)}",
    r'\\frac\{(.*?)\}\{(.*?)\}': lambda m: fr"{m.group(1)} over {m.group(2)}",
    r'\\sqrt\[(\\d+)\]\{(.*?)\}': lambda m: fr"the {m.group(1)} root of {m.group(2)}",
    r'\\sqrt\{(.*?)\}': lambda m: fr"the square root of {m.group(1)}",
    r'(\\operatorname\{.*?\}|\\\w+)': lambda m: m.group(1).replace('\\operatorname', '').replace('{', '').replace('}', ''),
    r'(\d+)\s*\*\s*(\d+)': lambda m: fr"{m.group(1)} times {m.group(2)}",
    r'(\d+)\s*\*\s*([a-zA-Z]+)': lambda m: fr"{m.group(1)} times {m.group(2)}",
    r'([a-zA-Z]+)\s*\*\s*(\d+)': lambda m: fr"{m.group(1)} times {m.group(2)}",
    r'([a-zA-Z]+)\^\{(\d+)\}': lambda m: fr"{m.group(1)} raised to the power of {m.group(2)}",
    r'(\d+)\^\{(\d+)\}': lambda m: fr"{m.group(1)} raised to the power of {m.group(2)}",
    r'\+': ' plus ',
    r'-': ' minus ',
    r'\*': ' times ',
    r'=': ' equals ',
    r'\\(\\d+)': lambda m: m.group(1),
    r'\\[a-zA-Z]+': lambda m: m.group(0).replace('\\', ''),
    r'([a-zA-Z]+)': lambda m: m.group(1),
    r'(\d+)': lambda m: m.group(1),
}


# Function to convert LaTeX expression to readable text using the conversion dictionary
def latex_to_readable(latex_expression):
    # Remove the surrounding $ symbols
    latex_expression = latex_expression.strip('$')
    
    # Apply conversions
    readable_text = latex_expression
    for pattern, replacement in conversion_dict.items():
        if callable(replacement):
            readable_text = re.sub(pattern, replacement, readable_text)
        else:
            readable_text = re.sub(pattern, replacement, readable_text)
    
    return readable_text.strip()

# Open the Excel workbook
file_path = 'nsmq_year_test.xlsx'  # Update with your actual file path
workbook = openpyxl.load_workbook(file_path)

# Iterate through each sheet in the workbook
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    if sheet_name.startswith('Round'):
        print(f'Processing sheet: {sheet_name}')
        
        # Iterate through column C
        for row in range(1, sheet.max_row + 1):
            cell = sheet[f'C{row}']
            if isinstance(cell.value, str) and cell.value.startswith('$') and cell.value.endswith('$'):
                latex_expression = cell.value
                readable_text = latex_to_readable(latex_expression)
                print(f'Row {row} LaTeX: {latex_expression} -> Readable: {readable_text}')
                cell.value = readable_text

# Save the modified workbook with a new name
modified_file_path = './modified_your_excel_file.xlsx'  # Update with your desired output file path
workbook.save(modified_file_path)

print(f'Modified file saved as {modified_file_path}')

