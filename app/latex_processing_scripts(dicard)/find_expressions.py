import openpyxl
import json
import re

# Load the Excel file
workbook = openpyxl.load_workbook('./nsmq_questions/2021/2021 NSMQ contest 1.xlsx')
sheet = workbook.active

# Define a function to find and return the location of expressions encapsulated with $
def find_expressions_in_column(sheet, column_index, word_symbol='$'):
    expression_locations = []
    pattern = re.compile(rf'\{word_symbol}(.*?)\{word_symbol}')

    for row_index in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=column_index).value
        if cell_value:
            matches = pattern.findall(cell_value)
            if matches:
                cell_location = sheet.cell(row=row_index, column=column_index).coordinate
                expression_locations.append({
                    "cell": cell_location,
                    "value": cell_value,
                    "expressions": matches,
                    "replacements": [""] * len(matches)  # Placeholder for replacements
                })
    return expression_locations

# Call the function with the column index where your sentences are stored
expression_locations = find_expressions_in_column(sheet, column_index=4)

# Save the expression locations to a JSON file
with open('expression_locations.json', 'w') as json_file:
    json.dump(expression_locations, json_file, indent=4)

print("Expression locations saved to 'expression_locations.json'")
