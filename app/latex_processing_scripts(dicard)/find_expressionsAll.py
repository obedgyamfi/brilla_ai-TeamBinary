import openpyxl
import json
import re

# Load the Excel file
workbook = openpyxl.load_workbook('./nsmq_questions/2021/2021 NSMQ contest 1.xlsx')

# Define a function to find and return the location of expressions encapsulated with $
def find_expressions_in_worksheets(workbook, word_symbol='$'):
    expression_locations = []
    pattern = re.compile(rf'\{word_symbol}(.*?)\{word_symbol}')

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row_index in range(1, sheet.max_row + 1):
            for col_index in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_index, column=col_index).value
                if cell_value is not None:  # Check if the cell value is not empty
                    cell_value = str(cell_value)  # Convert cell value to string
                    matches = pattern.findall(cell_value)
                    if matches:
                        cell_location = sheet.cell(row=row_index, column=col_index).coordinate
                        expression_locations.append({
                            "worksheet": sheet_name,
                            "cell": cell_location,
                            "value": cell_value,
                            "expressions": matches,
                            "replacements": [""] * len(matches)  # Placeholder for replacements
                        })
    return expression_locations


# Call the function to find expressions in all worksheets
expression_locations = find_expressions_in_worksheets(workbook)

# Save the expression locations to a JSON file
with open('expression_locations3.json', 'w') as json_file:
    json.dump(expression_locations, json_file, indent=4)

print("Expression locations saved to 'expression_locations.json'")
